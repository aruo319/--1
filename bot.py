import nextcord
from nextcord.ext import commands
import random
import requests, bs4, re
from datetime import datetime
import time
import difflib
import asyncio
import openpyxl
import asyncio, youtube_dl
import yt_dlp as youtube_dl
from datetime import timedelta
import asyncio
from datetime import datetime
import nextcord
from nextcord.ext import commands
from nextcord import Interaction, SlashOption, Embed
from nextcord.ext import commands, tasks
import nextcord
from nextcord.ext import commands
from nextcord import Interaction
from googletrans import Translator
from nextcord.ui import Button, View
import nextcord
from nextcord.ext import commands
from nextcord import Interaction, SlashOption
from nextcord.ext.commands.errors import CheckFailure
import nextcord
from nextcord.ext import commands
from nextcord import Interaction, SlashOption
import psutil

bot = commands.Bot(command_prefix=commands.when_mentioned_or("!") and "!", intents=nextcord.Intents.all())
                #  command_prefix 란 시작할 명령어

now = datetime.now()
Stime = f"{str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초"

now = datetime.now()


#"봇"이 준비 완료되면 터미널에 출력
@bot.event
async def on_ready():
    print(f'주의! 주의! 카사 발령보! 주의! 주의! {bot.user}')
    update_status.start()  # 봇이 준비되면 상태 업데이트 함수를 시작합니다

@tasks.loop(minutes=5)  # 5분마다 실행하도록 설정합니다
async def update_status():
    await bot.wait_until_ready()  # 봇이 완전히 준비될 때까지 기다립니다
    await bot.change_presence(activity=nextcord.Game(name=f"{len(bot.guilds)}개의 서버에서 활동"))
        # 사용자 지정 상태 설정법
        # status=nextcord.Status.online      (온라인)
        # status=nextcord.Status.idle        (자리 비움)
        # status=nextcord.Status.dnd         (다른 용무)
        # status=nextcord.Status.offline     (오프라인)
        #
        #   ~~하는 중 등 상태 설정법
        # activity=nextcord.Game(name="하는 중")
        # activity=nextcord.Streaming(name="방송 중", url="올리고 싶은 URL")
        # activity=nextcord.Activity(type=nextcord.ActivityType.listening, name="듣는 중")
        # activity=nextcord.Activity(type=nextcord.ActivityType.watching, name="시청 중")

async def main():
    async with bot:
        await bot.add_cog(Music(bot))
        await bot.start("MTI2OTE1OTUwMzM3NTYzNDQ3Mg.Gtog3Y.tMoLJNV6dfrgYvRJJXMSRu5QYZExnAeyxQ5Z74")


@bot.slash_command(name="타임아웃", description="선택한 유저를 타임아웃합니다.")
async def timeout_user(ctx: nextcord.Interaction,
                       멤버: nextcord.Member=nextcord.SlashOption(description="멤버를 입력하세요."),
                       시간: int=nextcord.SlashOption(description="시간을 입력하세요. (분 단위)")):
    
    if ctx.user.guild_permissions.administrator:  # 관리자 권한이 있는 경우에만 실행
        try:
            duration = timedelta(minutes=시간)  # 1분 타임아웃 설정
            await 멤버.timeout(duration, reason="슬래시 커맨드를 통한 타임아웃")
            await ctx.response.send_message(f"{멤버.mention}님이 {시간}분간 타임아웃 되었습니다.")
        except Exception as e:
            await ctx.response.send_message(f"타임아웃 중 오류가 발생했습니다: {e}")
    else:
        await ctx.response.send_message("이 명령어를 사용할 권한이 없습니다.", ephemeral=True)


@bot.slash_command(name="추방", description="유저를 추방함")
async def kick(ctx: nextcord.Interaction, 
               멤버: nextcord.Member = nextcord.SlashOption(description="추방할 멤버를 골라주세요.", required=True),
               사유: str = nextcord.SlashOption(description="사유를 적어주세요", required=False)):
    if ctx.user.guild_permissions.administrator:  # 관리자 권한이 있는 경우에만 실행
        try:
            await 멤버.kick(reason=사유)  # 추방코드
            await ctx.response.send_message(f'{멤버} 님이 추방되었습니다\n**사유** : {사유}')
        except Exception as e:
            await ctx.response.send_message(f"추방 중 오류가 발생했습니다: {e}")
    else:
        await ctx.response.send_message("이 명령어를 사용할 권한이 없습니다.", ephemeral=True)

@bot.slash_command(name="차단", description="유저를 차단함")
async def ban(ctx: nextcord.Interaction, 
              멤버: nextcord.Member = nextcord.SlashOption(description="차단할 멤버를 골라주세요.", required=True),
              사유: str = nextcord.SlashOption(description="사유를 적어주세요", required=False)):
    if ctx.user.guild_permissions.administrator:  # 관리자 권한이 있는 경우에만 실행
        try:
            await 멤버.ban(reason=사유)  # 차단코드
            await ctx.response.send_message(f'{멤버} 님이 차단되었습니다\n**사유** : {사유}')
        except Exception as e:
            await ctx.response.send_message(f"차단 중 오류가 발생했습니다: {e}")
    else:
        await ctx.response.send_message("이 명령어를 사용할 권한이 없습니다.", ephemeral=True)

#인사 명령어
@bot.command(name="안녕") # 명령
async def 인사(ctx):
    await ctx.send(f'{ctx.author.name}님 안녕하세요!')  # 답변

@bot.command(name="하이")  # 명령
async def 하이(ctx):
    await ctx.send(f'안녕하세요!')  # 답변


@bot.command(name="반응") # 명령
async def 인사(ctx):
    msg = await ctx.send(f'반응입니다!')  # 답변
    await msg.add_reaction("😁") # 추가하고 싶은 반응

@bot.event
async def on_reaction_add(reaction, user):
    if user.bot == 1: #봇이면 반응하지 않음
        return None
    # 이모지가 "😁" 이고 해당 메시지의 이모지인 경우에만 반응
    if str(reaction.emoji) == "😁" and reaction.message.content == '반응입니다!':
        await reaction.message.channel.send('이모지를 누르셨습니다!')  #눌렀을 때의 반응

@bot.slash_command(name="음식추천", description="카사가 음식을 추천해줘요") #name = 명령  description = 명령에 대한 설명
async def slash(ctx:nextcord.Interaction):
    ran = random.randint(0,4)  # 랜덤으로 보낼 답변의 갯수 4개라면 (0,3) 9개라면 (0,8)  [파이썬의 숫자는 0부터 시작]
    if ran == 0:  # 1번 랜덤
        r = "라면"   # 답변
    if ran == 1:  # 2번 랜덤
        r = "우동"   # 답변
    if ran == 2:
        r = "돈까스"
    if ran == 3:
        r = "김치 볶음밥"
    if ran == 4:
        r = "카레라이스"
    if ran == 5:
        r = "바나나 우유에 딸기 바른 식빵"
    await ctx.send(r, ephemeral=False)  # 변수 r의 값을 보냄

@bot.slash_command(name="자기소개", description="유저에게 자기소개를 합니다") #name = 명령  description = 명령에 대한 설명
async def slash(ctx:nextcord.Interaction):
    await ctx.send(f"안녕하세요!  저는 아루님이 2024년 7월 30일에 만들어주신 카사라고 해요!", ephemeral=False)  # 답변 , ephemeral= 이거는 메시지를 누구가 볼 수 있게 설정할지하는 코드
                                                            # True 이라면 메시지를 보낸이만 볼 수 있고
                                                            # flsae 이라면 모두가 볼 수 있음
@bot.slash_command(name="카사_초대링크", description="카사의 초대링크를 보여줍니다") #name = 명령  description = 명령에 대한 설명
async def slash(ctx:nextcord.Interaction):
    await ctx.send(f"https://discord.com/oauth2/authorize?client_id=1269159503375634472&permissions=8&integration_type=0&scope=bot", ephemeral=True)
    
                                                            # True 이라면 메시지를 보낸이만 볼 수 있고
                                                            # flsae 이라면 모두가 볼 수 있음

@bot.slash_command(name="날씨", description="날씨를 볼 수 있습니다")
async def slash(ctx: nextcord.Interaction, 지역:str=nextcord.SlashOption(description="지역을 입력하세요.")):
    await ctx.response.defer(ephemeral=False)
    try:
        # 네이버의 검색기능에서 지역을 추가하고 html정보를 받습니다
        검색 = 지역+"날씨"
        url = "https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=0&ie=utf8&query=" + 검색
        hdr = {'User-Agent': 'Mozilla/5.0'}
        req = requests.get(url, headers=hdr)
        html = req.text
        bsObj = bs4.BeautifulSoup(html, "html.parser")

        # 온도 정보를 가져옵니다
        온도 = bsObj.find('div', class_='temperature_text')
        온도텍 = 온도.text
        온도결과 = re.sub(r'[^0-9.]', '', 온도텍.strip().split('°')[0])
    
        # 체감온도 정보를 가져옵니다
        체감온도 = bsObj.find('div', class_='sort')
        체감온도텍 = 체감온도.text
        체감온도결과 = re.sub(r'[^0-9.]', '', 체감온도텍.strip().split('°')[0])

        # 미세먼지 결과를 가져오고 그에 따라 이모티콘을 넣습니다
        미세먼지 = bsObj.find('li', class_='item_today level2')
        미세2 = 미세먼지.find('span', class_='txt')
        미세먼지결과 = 미세2.text
        if 미세먼지결과=="좋음":
            미세결과 = "😀(좋음)"
        if 미세먼지결과=="보통":
            미세결과 = "😐(보통)"
        if 미세먼지결과=="나쁨":
            미세결과 = "😷(나쁨)"
        if 미세먼지결과=="매우나쁨":
            미세결과 = "😡(매우나쁨)"

        # 초미세먼지 결과를 가져오고 그에 따라 이모티콘을 넣습니다
        초미세먼지들 = bsObj.find_all('li', class_='item_today level2')
        if len(초미세먼지들) >= 2:
            초미세먼지 = 초미세먼지들[1]  # 두 번째로 나타나는 요소 선택
            미세2 = 초미세먼지.find('span', class_='txt')
            초미세먼지결과 = 미세2.text
        if 초미세먼지결과=="좋음":
            초미세결과 = "😀(좋음)"
        if 초미세먼지결과=="보통":
            초미세결과 = "😐(보통)"
        if 초미세먼지결과=="나쁨":
            초미세결과 = "😷(나쁨)"
        if 초미세먼지결과=="매우나쁨":
            초미세결과 = "😡(매우나쁨)"

        # 기후 정보를 가져옵니다
        기후 = bsObj.find('p', class_='summary')
        기후2 = 기후.find('span', class_='weather before_slash')
        기후결과 = 기후2.text
    


        # 가져온 정보들을 모두 임베드에 써놓습니다
        embed = nextcord.Embed(title=지역+' 날씨 정보',description='현재 온도',color=nextcord.Color(0x2ECCFA))
        embed.set_thumbnail(url="사진URL")
        embed.add_field(name=f"{온도결과}℃", value=f'체감 {체감온도결과}', inline=False)
        embed.add_field(name="미세먼지", value=f"{미세결과}", inline=False)
        embed.add_field(name="초미세먼지", value=f"{초미세결과}", inline=False)
        embed.add_field(name="기후", value=f"{기후결과}", inline=False)

        embed.set_footer(text=f"시각 : {str(now.hour)}시 {str(now.minute)}분 {str(now.second)}초")
    
        await ctx.send(embed=embed)

    # 에러가 발생할 시 보낼 메시지입니다
    except Exception as e:
        await ctx.send("올바른 지역을 입력해주세요")

@bot.slash_command(name="타수측정", description="타수를 측정합니다.") #name = 명령  description = 명령에 대한 설명
async def slash(ctx:nextcord.Interaction,주제=nextcord.SlashOption(choices=["메밀꽃 필 무렵", "동백꽃", "별 헤는 밤", "애국가", "금도끼", "별주부전", "님의 침묵"])):
    await ctx.send('타수 측정', ephemeral=True)

        
    if 주제 == "메밀꽃 필 무렵":
            def check(m):
                return m.author == ctx.user and m.channel == ctx.channel
        
            embed = nextcord.Embed(
                title='아래의 글을 입력하세요',

                description="주제 : 메밀꽃 필 무렵",
                color=nextcord.Color(0xFFFF00)
                )
            embed.add_field(name=f"**Enter하지 말고 그대로 써주세요(제한시간은 5분입니다)**",
                        value=f'```여름장이란 애시당초에 글러서 해는 아직 중천에 있건만 장판은 벌``````써 쓸쓸하고 더운 햇발이 벌려놓은 전시장 밑으로 등줄기를 훅훅 볶``````는다. 마을 사람들은 거의 돌아간 뒤요, 팔리지 못한 나무꾼 패가`````` 길거리에 궁싯거리고 들 있었으나, 석유 병이나 받고 고깃마리나`````` 사면 족할 것이```', inline=False)
        
            sentence = ["여름장이란 애시당초에 글러서 해는 아직 중천에 있건만 장판은 벌써 쓸쓸하고 더운 햇발이 벌려놓은 전시장 밑으로 등줄기를 훅훅 볶는다. 마을 사람들은 거의 돌아간 뒤요, 팔리지 못한 나무꾼 패가 길거리에 궁싯거리고 들 있었으나, 석유 병이나 받고 고깃마리나 사면 족할 것이"]

            await ctx.send("준비되면 1, 취소하려면 2를 입력하세요.", ephemeral=True)

            try:
                msg = await bot.wait_for("message", check=check, timeout=300)

                if msg.content == "1":
                    choice = random.choice(sentence)
                    await ctx.send(embed=embed, ephemeral=True)
                    startTime = time.time()
                    try:
                        answer = await bot.wait_for("message", check=check, timeout=300)

                        deltaTime = time.time() - startTime
                        accuracy = difflib.SequenceMatcher(None, choice, answer.content).ratio()
                        
                        타이핑한수 = len(answer.content)
                        deltaTime = time.time() - startTime
                        시간을분으로 = deltaTime / 60
                        일분타이핑수 = 타이핑한수 / 시간을분으로
                        정확합니다 = accuracy * 일분타이핑수
                        타수 = 정확합니다 / 5
                        레알타수 = 타수 * 10
                        레레알알타수 = round(레알타수)





                        if 레알타수 >= 1500:
                            await ctx.send('복붙금지입니다', ephemeral=True)
                        else:

                            
                            await ctx.send(f"**{ctx.user.name}님의 타수 : {레레알알타수}타\n정확도: {accuracy * 100:0.1f}%**", ephemeral=True)
                
                    except asyncio.exceptions.TimeoutError:
                        await ctx.send("시간이 초과되었습니다.", ephemeral=True)

                elif msg.content == "2":
                    await ctx.send("타자측정을 취소합니다.", ephemeral=True)

                else:
                    await ctx.send("올바른 값을 입력해주세요.", ephemeral=True)

            except asyncio.exceptions.TimeoutError:
                await ctx.send("시간이 초과되었습니다.\n**타수측정**을 입력해 다시 시도하세요.", ephemeral=True)

    if 주제 == "동백꽃":
            def check(m):
                return m.author == ctx.user and m.channel == ctx.channel
        
            embed = nextcord.Embed(
                title='아래의 글을 입력하세요',

                description="주제 : 동백꽃",
                color=nextcord.Color(0xFFFF00)
                )
            embed.add_field(name=f"**Enter하지 말고 그대로 써주세요(제한시간은 5분입니다)**",
                        value=f'```오늘도 또 우리 수탉이 막 쫓기었다. 내가 점심을 먹고 나무를 하러`````` 갈 양으로 나올 때이었다. 산으로 올라서려니까 등 뒤에서 푸드득``````푸드득 하고 닭의 횃소리가 야단이다. 깜짝 놀라서 고개를 돌려보니`````` 아니나 다르랴 두 놈이 또 얼리었다.```', inline=False)
        
            sentence = ["오늘도 또 우리 수탉이 막 쫓기었다. 내가 점심을 먹고 나무를 하러 갈 양으로 나올 때이었다. 산으로 올라서려니까 등 뒤에서 푸드득푸드득 하고 닭의 횃소리가 야단이다. 깜짝 놀라서 고개를 돌려보니 아니나 다르랴 두 놈이 또 얼리었다."]

            await ctx.send("준비되면 1, 취소하려면 2를 입력하세요.", ephemeral=True)

            try:
                msg = await bot.wait_for("message", check=check, timeout=300)

                if msg.content == "1":
                    choice = random.choice(sentence)
                    await ctx.send(embed=embed, ephemeral=True)
                    startTime = time.time()
                    try:
                        answer = await bot.wait_for("message", check=check, timeout=300)

                        deltaTime = time.time() - startTime
                        accuracy = difflib.SequenceMatcher(None, choice, answer.content).ratio()
                        
                        타이핑한수 = len(answer.content)
                        deltaTime = time.time() - startTime
                        시간을분으로 = deltaTime / 60
                        일분타이핑수 = 타이핑한수 / 시간을분으로
                        정확합니다 = accuracy * 일분타이핑수
                        타수 = 정확합니다 / 5
                        레알타수 = 타수 * 10
                        레레알알타수 = round(레알타수)





                        if 레알타수 >= 1500:
                            await ctx.send('복붙금지입니다', ephemeral=True)
                        else:

                            
                            await ctx.send(f"**{ctx.user.name}님의 타수 : {레레알알타수}타\n정확도: {accuracy * 100:0.1f}%**", ephemeral=True)
                
                    except asyncio.exceptions.TimeoutError:
                        await ctx.send("시간이 초과되었습니다.", ephemeral=True)

                elif msg.content == "2":
                    await ctx.send("타자측정을 취소합니다.", ephemeral=True)

                else:
                    await ctx.send("올바른 값을 입력해주세요.", ephemeral=True)

            except asyncio.exceptions.TimeoutError:
                await ctx.send("시간이 초과되었습니다.\n**타수측정**을 입력해 다시 시도하세요.", ephemeral=True)
        
    if 주제 == "별 헤는 밤":
            def check(m):
                return m.author == ctx.user and m.channel == ctx.channel
        
            embed = nextcord.Embed(
                title='아래의 글을 입력하세요',

                description="주제 : 별 헤는 밤",
                color=nextcord.Color(0xFFFF00)
                )
            embed.add_field(name=f"**Enter하지 말고 그대로 써주세요(제한시간은 5분입니다)**",
                        value=f'```계절이 지나가는 하늘에는 가을로 가득 차 있습니다. 나는 아무 걱``````정도 없이 가을 속의 별들을 다 헬 듯합니다. 가슴 속에 하나 둘 ``````새겨지는 별을 이제 다 못 헤는 것은 쉬이 아침이 오는 까닭이요, ``````내일 밤이 남은 까닭이요,```', inline=False)
        
            sentence = ["계절이 지나가는 하늘에는 가을로 가득 차 있습니다. 나는 아무 걱정도 없이 가을 속의 별들을 다 헬 듯합니다. 가슴 속에 하나 둘 새겨지는 별을 이제 다 못 헤는 것은 쉬이 아침이 오는 까닭이요, 내일 밤이 남은 까닭이요,"]

            await ctx.send("준비되면 1, 취소하려면 2를 입력하세요.", ephemeral=True)

            try:
                msg = await bot.wait_for("message", check=check, timeout=300)

                if msg.content == "1":
                    choice = random.choice(sentence)
                    await ctx.send(embed=embed, ephemeral=True)
                    startTime = time.time()
                    try:
                        answer = await bot.wait_for("message", check=check, timeout=300)

                        deltaTime = time.time() - startTime
                        accuracy = difflib.SequenceMatcher(None, choice, answer.content).ratio()
                        
                        타이핑한수 = len(answer.content)
                        deltaTime = time.time() - startTime
                        시간을분으로 = deltaTime / 60
                        일분타이핑수 = 타이핑한수 / 시간을분으로
                        정확합니다 = accuracy * 일분타이핑수
                        타수 = 정확합니다 / 5
                        레알타수 = 타수 * 10
                        레레알알타수 = round(레알타수)





                        if 레알타수 >= 1500:
                            await ctx.send('복붙금지입니다', ephemeral=True)
                        else:

                            
                            await ctx.send(f"**{ctx.user.name}님의 타수 : {레레알알타수}타\n정확도: {accuracy * 100:0.1f}%**", ephemeral=True)
                
                    except asyncio.exceptions.TimeoutError:
                        await ctx.send("시간이 초과되었습니다.", ephemeral=True)

                elif msg.content == "2":
                    await ctx.send("타자측정을 취소합니다.", ephemeral=True)

                else:
                    await ctx.send("올바른 값을 입력해주세요.", ephemeral=True)

            except asyncio.exceptions.TimeoutError:
                await ctx.send("시간이 초과되었습니다.\n**타수측정**을 입력해 다시 시도하세요.", ephemeral=True)
    if 주제 == "애국가":
            def check(m):
                return m.author == ctx.user and m.channel == ctx.channel
        
            embed = nextcord.Embed(
                title='아래의 글을 입력하세요',

                description="주제 : 애국가",
                color=nextcord.Color(0xFFFF00)
                )
            embed.add_field(name=f"**Enter하지 말고 그대로 써주세요(제한시간은 5분입니다)**",
                        value=f'```동해물과 백두산이 마르고 닳도록 하느님이 보우하사 우리나라 만세 ``````무궁화 삼천리 화려강산 대한사람 대한으로 길이 보전하세.```', inline=False)
        
            sentence = ["동해물과 백두산이 마르고 닳도록 하느님이 보우하사 우리나라 만세 무궁화 삼천리 화려강산 대한사람 대한으로 길이 보전하세."]

            await ctx.send("준비되면 1, 취소하려면 2를 입력하세요.", ephemeral=True)

            try:
                msg = await bot.wait_for("message", check=check, timeout=300)

                if msg.content == "1":
                    choice = random.choice(sentence)
                    await ctx.send(embed=embed, ephemeral=True)
                    startTime = time.time()
                    try:
                        answer = await bot.wait_for("message", check=check, timeout=300)

                        deltaTime = time.time() - startTime
                        accuracy = difflib.SequenceMatcher(None, choice, answer.content).ratio()
                        
                        타이핑한수 = len(answer.content)
                        deltaTime = time.time() - startTime
                        시간을분으로 = deltaTime / 60
                        일분타이핑수 = 타이핑한수 / 시간을분으로
                        정확합니다 = accuracy * 일분타이핑수
                        타수 = 정확합니다 / 5
                        레알타수 = 타수 * 10
                        레레알알타수 = round(레알타수)





                        if 레알타수 >= 1500:
                            await ctx.send('복붙금지입니다', ephemeral=True)
                        else:

                            
                            await ctx.send(f"**{ctx.user.name}님의 타수 : {레레알알타수}타\n정확도: {accuracy * 100:0.1f}%**", ephemeral=True)
                
                    except asyncio.exceptions.TimeoutError:
                        await ctx.send("시간이 초과되었습니다.", ephemeral=True)

                elif msg.content == "2":
                    await ctx.send("타자측정을 취소합니다.", ephemeral=True)

                else:
                    await ctx.send("올바른 값을 입력해주세요.", ephemeral=True)

            except asyncio.exceptions.TimeoutError:
                await ctx.send("시간이 초과되었습니다.\n**타수측정**을 입력해 다시 시도하세요.", ephemeral=True)


    if 주제 == "금도끼":
            def check(m):
                return m.author == ctx.user and m.channel == ctx.channel
        
            embed = nextcord.Embed(
                title='아래의 글을 입력하세요',

                description="주제 : 금도끼",
                color=nextcord.Color(0xFFFF00)
                )
            embed.add_field(name=f"**Enter하지 말고 그대로 써주세요(제한시간은 5분입니다)**",
                        value=f'```나무꾼 한 사람이 연못가에서 큰 나무를 베다가 번쩍 든 도끼를 놓``````쳐서 그 도끼가 연못물 속에 풍덩 들어가 버렸습니다. 한없이 깊은 ``````연못 속에 들어갔으니까 다시 찾을 생각도 못하고 나무꾼은 그냥 연``````못가에서 쓰려져서 탄식을 하고 있노라니까 어여쁜 물귀신이 나와서 ``````무엇 때```', inline=False)
        
            sentence = ["나무꾼 한 사람이 연못가에서 큰 나무를 베다가 번쩍 든 도끼를 놓쳐서 그 도끼가 연못물 속에 풍덩 들어가 버렸습니다. 한없이 깊은 연못 속에 들어갔으니까 다시 찾을 생각도 못하고 나무꾼은 그냥 연못가에서 쓰려져서 탄식을 하고 있노라니까 어여쁜 물귀신이 나와서 무엇 때"]

            await ctx.send("준비되면 1, 취소하려면 2를 입력하세요.", ephemeral=True)

            try:
                msg = await bot.wait_for("message", check=check, timeout=300)

                if msg.content == "1":
                    choice = random.choice(sentence)
                    await ctx.send(embed=embed, ephemeral=True)
                    startTime = time.time()
                    try:
                        answer = await bot.wait_for("message", check=check, timeout=300)

                        deltaTime = time.time() - startTime
                        accuracy = difflib.SequenceMatcher(None, choice, answer.content).ratio()
                        
                        타이핑한수 = len(answer.content)
                        deltaTime = time.time() - startTime
                        시간을분으로 = deltaTime / 60
                        일분타이핑수 = 타이핑한수 / 시간을분으로
                        정확합니다 = accuracy * 일분타이핑수
                        타수 = 정확합니다 / 5
                        레알타수 = 타수 * 10
                        레레알알타수 = round(레알타수)





                        if 레알타수 >= 1500:
                            await ctx.send('복붙금지입니다', ephemeral=True)
                        else:

                            
                            await ctx.send(f"**{ctx.user.name}님의 타수 : {레레알알타수}타\n정확도: {accuracy * 100:0.1f}%**", ephemeral=True)
                
                    except asyncio.exceptions.TimeoutError:
                        await ctx.send("시간이 초과되었습니다.", ephemeral=True)

                elif msg.content == "2":
                    await ctx.send("타자측정을 취소합니다.", ephemeral=True)

                else:
                    await ctx.send("올바른 값을 입력해주세요.", ephemeral=True)

            except asyncio.exceptions.TimeoutError:
                await ctx.send("시간이 초과되었습니다.\n**타수측정**을 입력해 다시 시도하세요.", ephemeral=True)


    if 주제 == "별주부전":
            def check(m):
                return m.author == ctx.user and m.channel == ctx.channel
        
            embed = nextcord.Embed(
                title='아래의 글을 입력하세요',

                description="주제 : 별주부전",
                color=nextcord.Color(0xFFFF00)
                )
            embed.add_field(name=f"**Enter하지 말고 그대로 써주세요(제한시간은 5분입니다)**",
                        value=f'```동해에 사는 용왕이 병이 들었는데, 좋다는 어떤 약도 소용이 없었``````다. 어느 날 도사가 나타나서 육지에 사는 토끼의 간을 먹으면 병이 ``````나을 것이라고 하였다. 용왕은 수궁의 대신들을 모아 놓고 육지에 ``````나갈 사자를 고르는데, 서로 다투기만 할 뿐 결정을 하지 못하였다.```', inline=False)
        
            sentence = ["동해에 사는 용왕이 병이 들었는데, 좋다는 어떤 약도 소용이 없었다. 어느 날 도사가 나타나서 육지에 사는 토끼의 간을 먹으면 병이 나을 것이라고 하였다. 용왕은 수궁의 대신들을 모아 놓고 육지에 나갈 사자를 고르는데, 서로 다투기만 할 뿐 결정을 하지 못하였다."]

            await ctx.send("준비되면 1, 취소하려면 2를 입력하세요.", ephemeral=True)

            try:
                msg = await bot.wait_for("message", check=check, timeout=300)

                if msg.content == "1":
                    choice = random.choice(sentence)
                    await ctx.send(embed=embed, ephemeral=True)
                    startTime = time.time()
                    try:
                        answer = await bot.wait_for("message", check=check, timeout=300)

                        deltaTime = time.time() - startTime
                        accuracy = difflib.SequenceMatcher(None, choice, answer.content).ratio()
                        
                        타이핑한수 = len(answer.content)
                        deltaTime = time.time() - startTime
                        시간을분으로 = deltaTime / 60
                        일분타이핑수 = 타이핑한수 / 시간을분으로
                        정확합니다 = accuracy * 일분타이핑수
                        타수 = 정확합니다 / 5
                        레알타수 = 타수 * 10
                        레레알알타수 = round(레알타수)





                        if 레알타수 >= 1500:
                            await ctx.send('복붙금지입니다', ephemeral=True)
                        else:

                            
                            await ctx.send(f"**{ctx.user.name}님의 타수 : {레레알알타수}타\n정확도: {accuracy * 100:0.1f}%**", ephemeral=True)
                
                    except asyncio.exceptions.TimeoutError:
                        await ctx.send("시간이 초과되었습니다.", ephemeral=True)

                elif msg.content == "2":
                    await ctx.send("타자측정을 취소합니다.", ephemeral=True)

                else:
                    await ctx.send("올바른 값을 입력해주세요.", ephemeral=True)

            except asyncio.exceptions.TimeoutError:
                await ctx.send("시간이 초과되었습니다.\n**타수측정**을 입력해 다시 시도하세요.", ephemeral=True)



    if 주제 == "님의 침묵":
            def check(m):
                return m.author == ctx.user and m.channel == ctx.channel
        
            embed = nextcord.Embed(
                title='아래의 글을 입력하세요',

                description="주제 : 님의 침묵",
                color=nextcord.Color(0xFFFF00)
                )
            embed.add_field(name=f"**Enter하지 말고 그대로 써주세요(제한시간은 5분입니다)**",
                        value=f'```님은 갔습니다. 아아, 사랑하는 나의 님은 갔습니다. 푸른 산빛을 ``````깨치고 단풍나무 숲을 향하여 난 작은 길을 걸어서, 차마 떨치고 갔``````습니다. 황금의 꽃갈이 굳고 빛나든 옛 맹서는 차디찬 티끌이 되어``````서 한숨의 미풍에 날아갔습니다.```', inline=False)
        
            sentence = ["님은 갔습니다. 아아, 사랑하는 나의 님은 갔습니다. 푸른 산빛을 깨치고 단풍나무 숲을 향하여 난 작은 길을 걸어서, 차마 떨치고 갔습니다. 황금의 꽃갈이 굳고 빛나든 옛 맹서는 차디찬 티끌이 되어서 한숨의 미풍에 날아갔습니다."]

            await ctx.send("준비되면 1, 취소하려면 2를 입력하세요.", ephemeral=True)

            try:
                msg = await bot.wait_for("message", check=check, timeout=300)

                if msg.content == "1":
                    choice = random.choice(sentence)
                    await ctx.send(embed=embed, ephemeral=True)
                    startTime = time.time()
                    try:
                        answer = await bot.wait_for("message", check=check, timeout=300)

                        deltaTime = time.time() - startTime
                        accuracy = difflib.SequenceMatcher(None, choice, answer.content).ratio()
                        
                        타이핑한수 = len(answer.content)
                        deltaTime = time.time() - startTime
                        시간을분으로 = deltaTime / 60
                        일분타이핑수 = 타이핑한수 / 시간을분으로
                        정확합니다 = accuracy * 일분타이핑수
                        타수 = 정확합니다 / 5
                        레알타수 = 타수 * 10
                        레레알알타수 = round(레알타수)




                        
                        if 레알타수 >= 1500:
                            await ctx.send('복붙금지입니다', ephemeral=True)
                        else:

                            
                            await ctx.send(f"**{ctx.user.name}님의 타수 : {레레알알타수}타\n정확도: {accuracy * 100:0.1f}%**", ephemeral=True)
                
                    except asyncio.exceptions.TimeoutError:
                        await ctx.send("시간이 초과되었습니다.", ephemeral=True)

                elif msg.content == "2":
                    await ctx.send("타자측정을 취소합니다.", ephemeral=True)

                else:
                    await ctx.send("올바른 값을 입력해주세요.", ephemeral=True)

            except asyncio.exceptions.TimeoutError:
                await ctx.send("시간이 초과되었습니다.\n**타수측정**을 입력해 다시 시도하세요.", ephemeral=True)

@bot.slash_command(name="가입",description="가입을 할 수 있습니다.")
async def 가입(ctx: nextcord.Interaction, 닉네임: str=nextcord.SlashOption(description="닉네임은 15글자까지 가능합니다.")):
    # 엑셀 파일 경로
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기 (없으면 새로 생성)
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # 새로운 유저 정보 추가
    user_id = str(ctx.user.id)

    # 이미 가입되어 있는지 확인
    for row in sheet.iter_rows(values_only=True):
        if row[0] == user_id:
            await ctx.send("이미 가입되어 있습니다.")
            return
        
    if len(닉네임) > 10: # 닉네임 제한
        await ctx.send("닉네임은 최대 10글자까지만 가능합니다.")
        return

    # 가입되어 있지 않으면 가입 처리
    row = [user_id, 닉네임]
    sheet.append(row)

    # 엑셀 파일 저장
    workbook.save(excel_file)

    await ctx.send(f'{닉네임}님, 가입이 완료되었습니다!')


@bot.slash_command(name="탈퇴",description="탈퇴을 할 수 있습니다.")
async def 탈퇴(ctx):

    # 엑셀 파일 경로
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # 유저의 디스코드 아이디 가져오기
        user_id = str(ctx.user.id)

        # 엑셀 파일에서 해당 유저 정보 찾기
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
            if row[0] == user_id:
                # 해당 유저 정보를 삭제하고 저장
                sheet.delete_rows(idx)
                workbook.save(excel_file)
                await ctx.send("탈퇴 처리되었습니다.")
                return
        
        # 만약 해당 유저 정보가 없는 경우
        await ctx.send("가입된 정보가 없습니다.")
        
    except FileNotFoundError:
        await ctx.send("가입된 정보가 없습니다.")




@bot.command(aliases=['입장'])
async def join(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel      # 입장코드
        await channel.connect()
        print("음성 채널 정보: {0.author.voice}".format(ctx))
        print("음성 채널 이름: {0.author.voice.channel}".format(ctx))
    else:
        embed = nextcord.Embed(title='음성 채널에 유저가 존재하지 않습니다.',  color=nextcord.Color(0xFF0000))
        await ctx.send(embed=embed)
 
@bot.command(aliases=['퇴장'])
async def out(ctx):
    try:
        await ctx.voice_client.disconnect()   #퇴장 코드
    except AttributeError as not_found_channel:
        embed = nextcord.Embed(title='봇이 존재하는 채널을 찾지 못하였습니다.',  color=nextcord.Color(0xFF0000))
        await ctx.send(embed=embed)




youtube_dl.utils.bug_reports_message = lambda: ''



ytdl_format_options = {
    'format': 'bestaudio/best',
    'outtmpl': '%(extractor)s-%(id)s-%(title)s.%(ext)s',
    'restrictfilenames': True,
    'noplaylist': True,
    'nocheckcertificate': True,
    'ignoreerrors': False,
    'logtostderr': False,
    'quiet': True,
    'no_warnings': True,
    'default_search': 'auto',
    'source_address': '0.0.0.0',  # bind to ipv4 since ipv6 addresses cause issues sometimes
}

ffmpeg_options = {
    'before_options': '-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5',
    'options': '-vn',
}

ytdl = youtube_dl.YoutubeDL(ytdl_format_options)

class YTDLSource(nextcord.PCMVolumeTransformer):
    def __init__(self, source, *, data, volume=0.5):
        super().__init__(source, volume)

        self.data = data

        self.title = data.get('title')
        self.url = data.get('url')

    @classmethod
    async def from_url(cls, url, *, loop=None, stream=False):
        loop = loop or asyncio.get_event_loop()
        data = await loop.run_in_executor(None, lambda: ytdl.extract_info(url, download=not stream))

        if 'entries' in data:
            # take first item from a playlist
            data = data['entries'][0]

        filename = data['url'] if stream else ytdl.prepare_filename(data)
        return cls(nextcord.FFmpegPCMAudio(filename, **ffmpeg_options), data=data)



class Music(commands.Cog):  #음악재생을 위한 클래스
    def __init__(self, bot):
        self.bot = bot



    @commands.command(aliases=['노래'])
    async def play(self, ctx, *, url):


        async with ctx.typing():
            player = await YTDLSource.from_url(url, loop=self.bot.loop, stream=True)
            ctx.voice_client.play(player, after=lambda e: print(f'플레이어 에러 : {e}') if e else None)
        embed = nextcord.Embed(title=f'현재 재생중인 음악 : {player.title}',  color=nextcord.Color(0xF3F781))
        await ctx.send(embed=embed)


    @commands.command(aliases=['볼륨'])
    async def volume(self, ctx, volume: int):


        if ctx.voice_client is None:
            embed = nextcord.Embed(title="음성 채널에 연결되지 않았습니다.",  color=nextcord.Color(0xFF0000))
            return await ctx.send(embed=embed)

        ctx.voice_client.source.volume = volume / 100  # 볼륨변경코드
        embed = nextcord.Embed(title=f"볼륨을 {volume}%으로 변경되었습니다.",  color=nextcord.Color(0x0040FF))
        await ctx.send(embed=embed)

    @commands.command(aliases=['삭제'])
    async def stop(self, ctx):


        await ctx.voice_client.disconnect()  # 음성채팅에서 나가는 코드

    @commands.command(aliases=['중지'])
    async def pause(self, ctx):


        if ctx.voice_client.is_paused() or not ctx.voice_client.is_playing():
            embed = nextcord.Embed(title="음악이 이미 일시 정지 중이거나 재생 중이지 않습니다.",  color=nextcord.Color(0xFF0000))
            await ctx.send(embed=embed)


        ctx.voice_client.pause()   # 정지하는 코드

    @commands.command(aliases=['재생'])
    async def resume(self, ctx):


        if ctx.voice_client.is_playing() or not ctx.voice_client.is_paused():   
            embed = nextcord.Embed(title="음악이 이미 재생 중이거나 재생할 음악이 존재하지 않습니다.",  color=nextcord.Color(0xFF0000))
            await ctx.send(embed=embed)

        ctx.voice_client.resume()    # 다시 재생하는 코드

    @play.before_invoke
    async def ensure_voice(self, ctx):
        if ctx.voice_client is None:
            if ctx.author.voice:
                await ctx.author.voice.channel.connect()
            else:
                embed = nextcord.Embed(title="음성 채널에 연결되어 있지 않습니다.",  color=nextcord.Color(0xFF0000))
                await ctx.send(embed=embed)
                raise commands.CommandError("작성자가 음성 채널에 연결되지 않았습니다.")
        elif ctx.voice_client.is_playing():
            ctx.voice_client.stop()

intents = nextcord.Intents.default()
intents.message_content = True

bot.add_cog(Music(bot))

@bot.command(name="카사바부")
async def greet(ctx):
    # 수정전에 나타낼 메시지
    msg = await ctx.send(f'ㅗ')
    
    # 2초간의 시간 후..
    await asyncio.sleep(0.05)

    # msg 변수를 안녕하세요로 수정함
    await msg.edit(content=f'너무해요 ㅠㅠ')

@bot.command(name="인사")
async def dm(ctx:nextcord.Intents):

    await ctx.author.send("안녕하세요! 저는 카사에요!")

@bot.slash_command(name="정보", description="유저의 정보를 불러옵니다")
async def user_info(ctx: nextcord.Interaction,
                    멤버: nextcord.Member = nextcord.SlashOption(description="정보를 알고 싶은 멤버를 입력하세요.", required=False)):
    
    if 멤버 == None:    # 만약 멤버를 선택하지 않았다면 멤버를 본인으로 설정
        멤버 = ctx.user
    
    embed = nextcord.Embed(
        title=f'**{멤버.display_name}**님의 정보',  # display_name는 사용자의 별명
        description=f'- {멤버}',
        color=nextcord.Color(0xD3851F)
    )
    embed.set_thumbnail(url=멤버.avatar.url)   
    # set_thumbnail를  .avatar.url을 사용하여 사용자의 프로필 링크로 설정

    embed.add_field(name=f'ID', value=f'{멤버.id}', inline=True)   # 멤버의 id

    bot_status = "🤖 **Bot**" if 멤버.bot else "👤 **User**"
    embed.add_field(name=f'Type', value=f'{bot_status}', inline=True)
    # 멤버가 봇이라면 🤖 **Bot** 유저라면 👤 **User** 봇,유저 구분은 멤버.bot을 이용

    embed.add_field(name=' ', value=' ', inline=False)  # 공백 필드 추가

    embed.add_field(name=f'가입 시기', value=f'{멤버.created_at}', inline=True) 
    embed.add_field(name=f'서버 가입 시기', value=f'{멤버.joined_at}', inline=True)
    # created_at는 디스코드 가입 시기이고 joined_at는 명령어를 입력한 서버에 가입한 날짜
    
    embed.add_field(name=' ', value=' ', inline=False)  # 공백 필드 추가

    role_mentions = [role.mention for role in 멤버.roles if role != ctx.guild.default_role]
    roles_str = ' '.join(role_mentions) if role_mentions else 'None'
    embed.add_field(name=f'보유 역할', value=f'{roles_str}', inline=True)
    # 멤버의 역할을 roles을 이용해서 추출하고 역할 중에서 != ctx.guild.default_role를 사용하여 에브리원 역할은 제외
    # 만약 보유한 역할이 없다면 None이라고 뜨게 함
    
    if 멤버.status == nextcord.Status.online:
        상태 = "🟢 온라인"
    elif 멤버.status == nextcord.Status.idle:
        상태 = "🌙 자리 비움"
    elif 멤버.status == nextcord.Status.dnd:
        상태 = "⛔ 방해 금지"
    else:
        상태 = "⚫ 오프라인"
    embed.add_field(name=f'상태', value=f'{상태}', inline=True)
    # 멤버의 status값을 추출하여 값에 따라 상태 변수에 저장

    user_status = 멤버.activity
    if user_status == None:
        pass
    else:
        embed.add_field(name=' ', value=' ', inline=False)  # 공백 필드 추가
        embed.add_field(name="상태메시지", value=user_status, inline=True)
    # 멤버의 activity값을 추출하여 상태메시지가 나오게 하고 상태메시지가 없다면 pass하여서 필드가 나타나지 않게 함
    
    await ctx.send(embed=embed) # 임베드 최종 추출
    
#id : 사용자의 ID (정수)
#name : 사용자의 이름
#discriminator : 사용자의 태그 번호
#display_name : 사용자 프로필의 별명
#mention : 사용자 멘션
#avatar : 사용자의 프로필
#avatar.url : 사용자의 프로필이 URL로 되있음
#bot : 를 이용하여 사용자가 봇인지 아닌지 구분 (True/False)
#system : 사용자가 디스코드 시스템 계정인지 구분 (True/False)
#guild : 현재 사용자가 속해있는 서버를 알려줌
#guild:name : 현재 사용자가 속해있는 서버의 이름을 알려줌
#nick : 서버 프로필의 별명
#joined_at : 사용자가 서버에 접속한 시각 (datetime 객체)
#created_at : 사용자가 디스코드에 가입한 시각 (datetime 객체)
#roles : 사용자의 역할 리스트
#top_role : 사용자가 가진 역할 중 가장 지휘가 높은 역할
#status : 사용자의 상태 (온라인, 자리 비움, 방해 금지, 오프라인)
#activity : 사용자의 상태 메시지
#voice : 사용자가 현재 속해있는 음성 채널
#guild_permissions : 서버에서 사용자가 가진 권한
#top_role.permissions : 사용자가 가진 가장 높은 역할의 권한
#premium_since : 사용자가 현재 속해있는 서버에서 부스트를 시작한 시간 (datetime 객체)
#pending : 사용자가 서버 규칙을 수락했는지 아닌지 (True/False)
#timed_out_until : 사용자가 타임아웃 된 시간 (datetime 객체)

@bot.command(name="메시지삭제")
async def delete_messages(ctx, num: int):
    if num < 1:
        await ctx.send("1 이상의 숫자를 입력하세요.")
        return
    
    # 사용자로부터 입력받은 숫자만큼 메시지를 삭제합니다.
    await ctx.channel.purge(limit=num + 1 )#명령어 메시지도 포함하여 삭제
    await ctx.send(f'메시지를 삭제했습니다.', delete_after=2)

@bot.command(name="갤갤갤갤")
async def emojis(ctx: nextcord.Interaction):
    emoji = bot.get_emoji(1163769571401023511)

    await ctx.send(f"갤갤갤갤갤  {emoji}")

@bot.slash_command(name="라디오", description="라디오를 재생합니다.")
async def radio(interaction: Interaction, station_name: str = SlashOption(description="라디오 방송국 이름을 입력하세요")):
    if not interaction.user.voice or not interaction.user.voice.channel:
        await interaction.response.send_message("먼저 음성 채널에 들어가야 합니다.", ephemeral=True)
        return

    channel = interaction.user.voice.channel
    voice_client = nextcord.utils.get(bot.voice_clients, guild=interaction.guild)

    if voice_client is None:
        voice_client = await channel.connect()
#api
    response = requests.get(f'http://de1.api.radio-browser.info/json/stations/byname/{station_name}')
    stations = response.json()

    if not stations:
        await interaction.response.send_message("해당 이름의 라디오 방송국을 찾을 수 없습니다.", ephemeral=True)
        return

    station = stations[0]
    stream_url = station['url']
    station_name = station['name']
    station_frequency = station.get('frequency', 'N/A')

    if voice_client.is_playing():
        voice_client.stop()

    voice_client.play(nextcord.FFmpegPCMAudio(stream_url), after=lambda e: print(f'Error: {e}') if e else None)

    embed = nextcord.Embed(title="라디오 재생", description=f"{station_name} 방송을 재생합니다.", color=0x00ff00)
    embed.add_field(name="방송국 이름", value=station_name, inline=True)
    embed.add_field(name="주파수", value=station_frequency, inline=True)
    embed.add_field(name="스트림 URL", value=stream_url, inline=False)

    await interaction.response.send_message(embed=embed)

@bot.slash_command(name="stop", description="라디오 스트림을 중지합니다.")
async def stop(interaction: Interaction):
    voice_client = nextcord.utils.get(bot.voice_clients, guild=interaction.guild)
    if voice_client and voice_client.is_playing():
        voice_client.stop()
        await voice_client.disconnect()
        await interaction.response.send_message("라디오 스트림을 중지했습니다.")
    else:
        await interaction.response.send_message("현재 재생 중인 라디오가 없습니다.", ephemeral=True)
        
@bot.slash_command(name="돈받기", description="5000원을 받을 수 있습니다.")
async def 하이(ctx):
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # 입력한 사용자의 아이디
        user_id = str(ctx.user.id)

        # 사용자 아이디가 있는 행 인덱스 찾기
        target_row_index = None

        # 첫 번째 열을 순회하며 사용자 아이디가 있는 행 인덱스 찾기
        for row_index in range(1, sheet.max_row + 1):
            for cell in sheet.iter_cols(min_row=row_index, max_row=row_index, min_col=1, max_col=1, values_only=True):
                if cell[0] == user_id:
                    target_row_index = row_index
                    break
            if target_row_index is not None:
                break  # 찾았으므로 더 이상 반복하지 않음

        # 사용자 아이디가 있는 행 인덱스 출력
        if target_row_index is None:
            await ctx.send("가입을 해주세요.")  # 행에 아이디가 존재하지 않을 때
            return

        # 마지막으로 돈을 받은 날짜 확인
        last_claim_date = sheet.cell(row=target_row_index, column=4).value
        if last_claim_date is not None:
            last_claim_date = datetime.strptime(last_claim_date, '%Y-%m-%d')
            if last_claim_date.date() == datetime.now().date():
                await ctx.send("오늘은 돈받기를 쓰셨네요! 다음날에 또 와서 받아주세요!"
                               )
                return

        current_value = sheet.cell(row=target_row_index, column=3).value
        if current_value is None:
            current_value = 0
        current_value = int(current_value)

        # 새로운 값 계산
        new_value = current_value + 5000  # 기존 값에 5000을 더함

        # 값 업데이트
        sheet.cell(row=target_row_index, column=3).value = new_value
        sheet.cell(row=target_row_index, column=4).value = datetime.now().strftime('%Y-%m-%d')  # 마지막으로 돈을 받은 날짜 업데이트

        # 엑셀 파일 저장
        workbook.save(excel_file)

        embed = nextcord.Embed(
            title=f'{ctx.user.name}',  # 제목과 설명은 임베드에 1개만 추가가 가능합니다
            description='돈 잔액',
            color=nextcord.Color(0xF3F781)  # 색상 코드
        )
        embed.add_field(name='추가된 잔액', value='5000원', inline=False)  # 추가된 돈을 보여줌
        embed.add_field(name='현재 잔액', value=f'{new_value}', inline=False)  # 자신의 잔액을 보여줌
        await ctx.send(embed=embed, ephemeral=False)

    except FileNotFoundError:
        print(f"파일 '{excel_file}'을(를) 찾을 수 없습니다.")
    except Exception as e:
        print(f"에러 발생: {e}")
        
@bot.slash_command(name="잔액", description="잔액을 알려줍니다.")
async def slash(ctx):
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # 입력한 사용자의 아이디
        user_id = str(ctx.user.id)

        # 사용자 아이디가 있는 행 인덱스 찾기
        target_row_index = None

        # 첫 번째 열을 순회하며 사용자 아이디가 있는 행 인덱스 찾기
        for row_index in range(1, sheet.max_row + 1):
            for cell in sheet.iter_cols(min_row=row_index, max_row=row_index, min_col=1, max_col=1, values_only=True):
                if cell[0] == user_id:
                    target_row_index = row_index
                    break
            if target_row_index is not None:
                break  # 찾았으므로 더 이상 반복하지 않음

        # 사용자 아이디가 있는 행 인덱스 출력
        if target_row_index is None:
            await ctx.send("가입을 해주세요.")
            return

        current_value = sheet.cell(row=target_row_index, column=3).value

        current_value = int(current_value)


        # 엑셀 파일 저장
        workbook.save(excel_file)
        embed = nextcord.Embed(
            title=f'{ctx.user.name}',           # 제목과 설명은 임베드에 1개만 추가가 가능합니다
            description='돈 잔액',
            color=nextcord.Color(0xF3F781)  # 색상 코드
        )
        embed.add_field(name='현재 잔액', value=f'{current_value}', inline=False) # 필드
        
        
        await ctx.send(embed=embed, ephemeral=False)


    except FileNotFoundError:
        print(f"파일 '{excel_file}'을(를) 찾을 수 없습니다.")
    except Exception as e:
        print(f"에러 발생: {e}")

@bot.slash_command(name="출석체크", description="출석체크를 합니다.")
async def 출석체크(ctx):
    excel_file = 'data.xlsx'

    try:
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # 입력한 사용자의 아이디
        user_id = str(ctx.user.id)

        # 사용자 아이디가 있는 행 인덱스 찾기
        target_row_index = None

        # 첫 번째 열을 순회하며 사용자 아이디가 있는 행 인덱스 찾기
        for row_index in range(1, sheet.max_row + 1):
            for cell in sheet.iter_cols(min_row=row_index, max_row=row_index, min_col=1, max_col=1, values_only=True):
                if cell[0] == user_id:
                    target_row_index = row_index
                    break
            if target_row_index is not None:
                break  # 찾았으므로 더 이상 반복하지 않음

        # 사용자 아이디가 있는 행 인덱스 출력
        if target_row_index is None:
            await ctx.send("가입을 해주세요.")  # 행에 아이디가 존재하지 않을 때
            return

        # 마지막으로 출석체크를 한 날짜 확인
        last_check_date = sheet.cell(row=target_row_index, column=4).value
        if last_check_date is not None:
            last_check_date = datetime.strptime(last_check_date, '%Y-%m-%d')
            if last_check_date.date() == datetime.now().date():
                await ctx.send("오늘은 이미 출석체크를 하셨습니다! 내일 다시 시도해 주세요.")
                return

        current_value = sheet.cell(row=target_row_index, column=3).value
        if current_value is None:
            current_value = 0
        current_value = int(current_value)

        # 새로운 값 계산
        new_value = current_value + 5000  # 기존 값에 5000을 더함

        # 값 업데이트
        sheet.cell(row=target_row_index, column=3).value = new_value
        sheet.cell(row=target_row_index, column=4).value = datetime.now().strftime('%Y-%m-%d')  # 마지막으로 출석체크를 한 날짜 업데이트

        # 엑셀 파일 저장
        workbook.save(excel_file)

        embed = nextcord.Embed(
            title=f'{ctx.user.name}',  # 제목과 설명은 임베드에 1개만 추가가 가능합니다
            description='출석체크 완료!',
            color=nextcord.Color(0xF3F781)  # 색상 코드
        )
        embed.add_field(name='추가된 잔액', value='5000원', inline=False)  # 추가된 돈을 보여줌
        embed.add_field(name='현재 잔액', value=f'{new_value}', inline=False)  # 자신의 잔액을 보여줌
        await ctx.send(embed=embed, ephemeral=False)

    except FileNotFoundError:
        print(f"파일 '{excel_file}'을(를) 찾을 수 없습니다.")
    except Exception as e:
        print(f"에러 발생: {e}")
        
class CalculatorView(View):
    def __init__(self):
        super().__init__()
        self.expression = ""
        self.last_button = ""

    @nextcord.ui.button(label='1', style=nextcord.ButtonStyle.secondary)
    async def one(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '1'
        self.last_button = '1'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='2', style=nextcord.ButtonStyle.secondary)
    async def two(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '2'
        self.last_button = '2'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='3', style=nextcord.ButtonStyle.secondary)
    async def three(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '3'
        self.last_button = '3'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='+', style=nextcord.ButtonStyle.success)
    async def add(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '+'
        self.last_button = '+'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='⌫', style=nextcord.ButtonStyle.danger)
    async def backspace(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression = self.expression[:-1]
        self.last_button = '⌫'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='4', style=nextcord.ButtonStyle.secondary)
    async def four(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '4'
        self.last_button = '4'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='5', style=nextcord.ButtonStyle.secondary)
    async def five(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '5'
        self.last_button = '5'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='6', style=nextcord.ButtonStyle.secondary)
    async def six(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '6'
        self.last_button = '6'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='-', style=nextcord.ButtonStyle.success)
    async def subtract(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '-'
        self.last_button = '-'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='C', style=nextcord.ButtonStyle.danger)
    async def clear(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression = ''
        self.last_button = ''
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='7', style=nextcord.ButtonStyle.secondary)
    async def seven(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '7'
        self.last_button = '7'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='8', style=nextcord.ButtonStyle.secondary)
    async def eight(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '8'
        self.last_button = '8'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='9', style=nextcord.ButtonStyle.secondary)
    async def nine(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '9'
        self.last_button = '9'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='×', style=nextcord.ButtonStyle.success)
    async def multiply(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '*'
        self.last_button = '*'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='ㅤ', style=nextcord.ButtonStyle.danger)
    async def nothing(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        pass  

    @nextcord.ui.button(label='00', style=nextcord.ButtonStyle.secondary)
    async def zerozero(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        if self.expression == "" or self.expression[-1] in "+-*/":
            return  
        self.expression += '00'
        self.last_button = '00'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='0', style=nextcord.ButtonStyle.secondary)
    async def zero(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '0'
        self.last_button = '0'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='.', style=nextcord.ButtonStyle.secondary)
    async def dot(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        if self.expression and self.expression[-1] in '0123456789':
            self.expression += '.'
        self.last_button = '.'
        await interaction.response.edit_message(content=self.expression)    

    @nextcord.ui.button(label='÷', style=nextcord.ButtonStyle.primary)
    async def divide(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        self.expression += '/'
        self.last_button = '/'
        await interaction.response.edit_message(content=self.expression)

    @nextcord.ui.button(label='=', style=nextcord.ButtonStyle.primary)
    async def equals(self, button: nextcord.ui.Button, interaction: nextcord.Interaction):
        try:
            self.expression = str(eval(self.expression.replace('×', '*').replace('÷', '/')))
        except Exception:
            self.expression = '오류가 났어요! `C`버튼을 눌러 다시 식을 써주세요!'
        self.last_button = '='
        await interaction.response.edit_message(content=self.expression)

@bot.slash_command(name='계산기', description='계산기를 엽니다')
async def calculator(interaction: Interaction):
    view = CalculatorView()
    embed = nextcord.Embed(title="계산기", description="> 아래 버튼들을 활용하여 계산기 기능을 써주세요!\n > 다른 식을 적고 싶으시면 `C`버튼을 눌러주세요", color=0x00ff00)
    await interaction.response.send_message(embed=embed, view=view)
    
@bot.slash_command(name="음악봇사용법", description="음악봇의 명령어를 정리해놨습니다.") 
async def slash(ctx:nextcord.Interaction):
    await ctx.send(f"`!입장`은 자신이 있는 보이스룸에 봇이 입장합니다. \n `!퇴장`은 봇이 퇴장합니다. \n `!노래`는 현재 재생중인 노래를 표시합니다. \n `!중지`는 음악을 중지시킵니다. \n `!재생`은 다시 음악을 재생시킵니다. \n 아직 음악봇의 기능이 적어서 대기열 기능이 있긴 하지만, 임베드에 표시되지 않아서 기존 곡이 끝나면 자동으로 실행됩니다!", ephemeral=False) 
    
@bot.slash_command(name='청소', description='지정한 개수만큼 메시지를 삭제합니다')
@commands.has_permissions(administrator=True)
async def clean_messages(
    interaction: Interaction, 
    count: int = SlashOption(description="삭제할 메시지의 수를 입력하세요", required=True, min_value=1, max_value=100)
):
    if not interaction.channel:
        await interaction.response.send_message("이 명령어는 채널에서만 사용할 수 있습니다.", ephemeral=True)
        return
    
    try:
        deleted = await interaction.channel.purge(limit=count + 1)
        await interaction.response.send_message(f'{len(deleted) - 1}개의 메시지를 삭제했습니다.', ephemeral=True)
    except Exception as e:
        await interaction.response.send_message(f"메시지 삭제 중 오류가 발생했습니다: {str(e)}", ephemeral=True)
 
@bot.slash_command(name="사용량", description="서버의 CPU 및 RAM 사용량을 확인합니다")
async def usage(interaction: Interaction):
    # CPU 사용량
    cpu_usage = psutil.cpu_percent(interval=1)
    
    # RAM 사용량
    memory_info = psutil.virtual_memory()
    total_memory = memory_info.total / (1024 ** 3)  # GB 단위로 변환
    used_memory = memory_info.used / (1024 ** 3)  # GB 단위로 변환
    available_memory = memory_info.available / (1024 ** 3)  # GB 단위로 변환

    # 결과 메시지 생성
    result_message = (
        f"**CPU 사용량**: {cpu_usage}%\n"
        f"**전체 RAM**: {total_memory:.2f} GB\n"
        f"**사용된 RAM**: {used_memory:.2f} GB\n"
        f"**사용 가능한 RAM**: {available_memory:.2f} GB"
    )

    # 결과 메시지 전송
    await interaction.response.send_message(result_message)
    
@bot.command(name="뽑기")
async def tokenboopgi(ctx):
    if random.randint(1, 100000000) == 1:
        await ctx.send("MTI2ODU0OTcyNTDFKSDDA0NDU5Ng.GFa1NU.DKFOSDKFMLSKDJ0QNyajirhduwIT4<= 이것은 카사의 토큰!" + " 속았었죠? ㅎㅎ")
    else:
        await ctx.send("저런~ 안타깝게도 카사 토큰을 뽑지 못하셨네요~ 아쉽네요~")
    
bot.run("토큰") #토큰]

